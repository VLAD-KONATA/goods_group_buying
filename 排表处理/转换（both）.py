import pandas as pd
import os
from openpyxl import load_workbook
import numpy as np
#=IF(C4="",B4,CONCAT(B4,"（",C4,"）"))
#如果盲抽cn和种类为两列的excel处理方式


def shopping(name,file_path,filename,obj_path,wb,sheetname,storage):
    print("{}/{} processing.\n".format(name,sheetname))
    sheet=wb[sheetname] #遍历sheets并打开对应sheet
    max_col = sheet.max_column  #获取最大列数
    '''
    sheetname=pd.read_excel(file_path,header=None,nrows=1,sheet_name=sheetname)
    sheetname=sheetname.iloc[0,0].split('【')[-1].split('】')[0]
    '''
    table = pd.read_excel(
    file_path,
    skiprows=1,        # 跳过第一行
    sheet_name=sheetname
    )
    
    try:
        oall=table.count()['种类']
    except:
        oall=table.count()['谷名']
    
    #统计扫街部分
    shopping_list=[]
    nall=0
    for index,contents in table.iterrows():
        if pd.isna(contents['cn']):
            pass
        else:
            cn=contents['cn']
        try:
            goodsname=contents['种类']
        except:
            goodsname=contents['谷名']
            
        shopping_list.append([cn,filename,sheetname,goodsname,1,'',storage,'不可排发'])
        nall+=1
    shopping_list=pd.DataFrame(shopping_list,columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #重复项合并求和
    shopping_list=shopping_list.groupby(['cn','系列','表名','谷名','肾/国际状态','囤货','排发状态'],as_index=False).sum().reindex(columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #新建系列文件夹
    if not os.path.exists(obj_path):
        os.makedirs(obj_path,exist_ok=True)
    #按照sheetname分别存储
    shopping_list.to_excel(os.path.join(obj_path,f'{sheetname}.xlsx'),sheet_name=sheetname,header=True,index=False)
    print("{}/{} completed {}/{}.\n".format(name,sheetname,oall,nall))

def schedule(name,file_path,filename,obj_path,keywords,wb,sheetname,storage):
    print("{}/{} processing.\n".format(name,sheetname))
    #生成筛选盲抽部分的正则表达式keyline
    for i in range(len(keywords)):
        if i==0:
            keyline=keywords[i]
        else:
            keyline+='|'+keywords[i]    
        
    sheet=wb[sheetname] #遍历sheets并打开对应sheet
    max_col = sheet.max_column  #获取最大列数
    '''
    sheetname=pd.read_excel(file_path,header=None,nrows=1,sheet_name=sheetname)
    sheetname=sheetname.iloc[0,0].split('【')[-1].split('】')[0]
    '''
    table = pd.read_excel(
    file_path,
    skiprows=1,        # 跳过第一行
    usecols=range(1,max_col),     # 从 B 列开始读取（跳过第一列）
    sheet_name=sheetname
    ).drop(0)   #读取表格
    oall=table.count().sum()
    #通过正则筛选盲抽/非盲抽部分
    table_blind=table.loc[:, table.columns.str.contains(keyline,regex=True)]    
    table_unblind = table.loc[:, ~table.columns.str.contains(keyline,regex=True)]
    
    #统计非盲抽部分
    unblind_list=[]
    unblinds=0
    for index,content in table_unblind.items():
        goodsname=index
        for cn in content.values:
            if pd.isna(cn):
                continue
            else:
                unblind_list.append([cn,filename,sheetname,goodsname,1,'',storage,'不可排发'])
                unblinds+=1
    unblind_list=pd.DataFrame(unblind_list,columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #重复项合并求和
    unblind_list=unblind_list.groupby(['cn','系列','表名','谷名','肾/国际状态','囤货','排发状态'],as_index=False).sum().reindex(columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    
    #统计盲抽部分
    blind_list=[]
    blinds=0
    for index,contents in table_blind.items():
        goodsname=index
        for content in contents.values:
            if pd.isna(content):
                continue
            else:
                if "（" not in str(content) and "(" not in str(content):  #如果盲抽没有打开确认
                    cn=content
                    blind_list.append([cn,filename,sheetname,goodsname,1,'',storage,'不可排发'])
                else:
                    if "(" in content: #如果前括号是英文括号
                        cn=content.split('(')[0]
                        if ")" in content:  #如果后括号是英文括号
                            chara=content.split('(')[-1].split(')')[0]
                        elif "）" in content:  #如果后括号是中文括号
                            chara=content.split('(')[-1].split('）')[0]
                    elif "（" in content:   #如果前括号是中文括号
                        cn=content.split('（')[0]
                        if ")" in content:  #如果后括号是英文括号
                            chara=content.split('（')[-1].split(')')[0]
                        elif "）" in content:  #如果后括号是中文括号
                            chara=content.split('（')[-1].split('）')[0]
                    blind_list.append([cn,filename,sheetname,goodsname+chara,1,'',storage,'不可排发'])
                blinds+=1
    blind_list=pd.DataFrame(blind_list,columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #重复项合并求和
    blind_list=blind_list.groupby(['cn','系列','表名','谷名','肾/国际状态','囤货','排发状态'],as_index=False).sum().reindex(columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #合并盲抽/非盲抽部分
    final_list=pd.concat([blind_list,unblind_list], axis=0, ignore_index=True)
    nall=unblinds+blinds
    #新建系列文件夹
    if not os.path.exists(obj_path):
        os.makedirs(obj_path,exist_ok=True)
    #按照sheetname分别存储
    final_list.to_excel(os.path.join(obj_path,f'{sheetname}.xlsx'),sheet_name=sheetname,header=True,index=False)
    print("{}/{} completed {}/{}.\n".format(name,sheetname,oall,nall))
    
    
    
name='1月、2月、3月拼煤'
#name='台湾30周年展+快闪+逃脱（已到齐）'

#筛选盲抽部分的正则表达式关键词
keywords=["盲抽","扭蛋","随机","特典","单抽","闪","自选"]
#keywords=["盲抽","land方3","OPED吧唧扭蛋","拼图盲","光栅卡","立牌","贴纸","app挂件","百鬼Q","投球游戏","摇曳的亚克力支架","app"]
#keywords=["盲抽","造型收藏卡","军牌"]
#keywords=["柯公式服","平公式服","和公式服","k捧花","新公式服","兰公式服","红"]
#keywords=["盲抽","land方3","OPED吧唧扭蛋","拼图盲","光栅卡","钥匙孔盲抽1","钥匙孔盲抽2","钥匙孔盲抽4","立牌","app","app挂件","百鬼Q吧唧","贴纸","投球游戏","摇曳的亚克力支架"]
#keywords=["盲抽","随机","扭蛋","饮品卡","台北","高雄","台中"]

#设置囤货
storages=['萌','dock','黄油']
storage=storages[1]

#设置主要格式化模式为扫街表/排表
mainmodes=['schedule','shopping']
mainmode=mainmodes[0]

#主模式为排表时sheet里的扫街表请添加在shoppingsheets中
shoppingsheets=["泰国吧唧选配"]
#shoppingsheets=["7.8台湾快闪单领","台湾展快闪单领","8.13台湾快闪扫街","12.13台湾扫街单领","6月27日限定相卡打印","6月27日单领"]

#主模式为扫街表时sheet里的排表请添加在schedulesheets中
schedulesheets=["杂志便签"]

#更新特定新分表/旧分表用
update=False
updatesheets=["ベツコミ5月号小卡（包国际不包国内，加单不保，捆序 松 研二）"]

file_path = rf"./排发表/{storage}/{name}.xlsx"

filename=os.path.basename(file_path).split('.xlsx')[0]  #获取文件名/系列
obj_path=f'./导入表/{storage}/{filename}'   #生成目标地址

wb=load_workbook(file_path)

for sheetname in wb.sheetnames:
    if (update and sheetname in updatesheets) or (not update):
        if mainmode=='shopping':
            if sheetname in schedulesheets:
                schedule(name,file_path,filename,obj_path,keywords,wb,sheetname,storage)
            else:
                shopping(name,file_path,filename,obj_path,wb,sheetname,storage)
        elif mainmode=='schedule':
            if sheetname in shoppingsheets:
                shopping(name,file_path,filename,obj_path,wb,sheetname,storage)
            else:
                schedule(name,file_path,filename,obj_path,keywords,wb,sheetname,storage)
    else:
        continue
    
                