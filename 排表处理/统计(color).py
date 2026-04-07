from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
import pandas as pd
import os

def get_cell_colors(wb,cns,sheetname):
    """
    统计Excel中所有内容为"dock"的单元格背景色
    
    参数:
        file_path: Excel文件路径
        sheet_name: 工作表名称(默认为"雪夜Q吧唧")
        
    返回:
        字典: {背景色RGB值: 出现次数}
    """
    # 加载工作簿和工作表
    sheet=wb[sheetname] #遍历sheets并打开对应sheet
    colors = []

    for cn in cns:
        # 遍历所有单元格
        for row in sheet.iter_rows():
            for cell in row:
                # 检查单元格值是否为"dock"
                if str(cell.value) == str(cn):
                    fill = cell.fill
                    

                    # 获取前景色(通常代表填充色)
                    color = fill.fgColor.rgb if fill.fgColor else None                        
                    
                    # 统计颜色
                    if color:
                        colors.append([cn,str(color),1])
                    else:
                        # 无填充色的情况
                        colors.append([cn,'无填充色',1])
    colors=pd.DataFrame(colors,columns=['cn','颜色','数量'])
    #重复项合并求和
    colors=colors.groupby(['cn','颜色'],as_index=False).sum().reindex(columns=['cn','颜色','数量'])
    #合并盲抽/非盲抽部分
    return colors

def save_or_append_to_excel(df, obj_path, sheet_name):
    """
    将DataFrame保存到Excel文件
    - 如果文件不存在，创建新文件
    - 如果文件存在，追加新Sheet（如果Sheet已存在则覆盖）
    
    参数:
        df: 要保存的DataFrame
        file_path: Excel文件路径
        sheet_name: 要写入的Sheet名称
        index: 是否保留索引（默认False）
    """
    # 检查文件是否存在
    if not os.path.exists(obj_path):
        # 文件不存在，创建新文件
        with pd.ExcelWriter(obj_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, header=True,index=False)
        print(f"创建新文件 {obj_path} 并写入Sheet '{sheet_name}'")
    else:
        # 文件存在，加载工作簿
        book = load_workbook(obj_path)
        
        # 检查Sheet是否已存在
        if sheet_name in book.sheetnames:
            # 删除已有Sheet
            std = book[sheet_name]
            book.remove(std)
            print(f"Sheet '{sheet_name}' 已存在，将被覆盖")
        
        # 创建新的ExcelWriter对象
        with pd.ExcelWriter(obj_path, engine='openpyxl') as writer:
            # 将已存在的Sheet写入writer
            writer.book = book
            # 将其他Sheet复制到writer
            for sheet in book.sheetnames:
                if sheet != sheet_name:  # 跳过我们要写入的Sheet
                    writer.sheets[sheet] = book[sheet]
            # 写入新数据
            df.to_excel(writer, sheet_name=sheet_name,header=True,index=False)
        
        print(f"在现有文件 {obj_path} 中追加/更新Sheet '{sheet_name}'")
        
def get_cns_schedule(name,file_path,wb,sheetname):
    print("{}/{} processing.\n".format(name,sheetname))
    sheet=wb[sheetname] #遍历sheets并打开对应sheet
    max_col = sheet.max_column  #获取最大列数
    '''
    sheetname=pd.read_excel(file_path,header=None,nrows=1,sheet_name=sheetname)
    sheetname=sheetname.iloc[0,0].split('【')[-1].split('】')[0]
    '''
    table = pd.read_excel(
    file_path,
    skiprows=1,        # 跳过第一行
    usecols=range(1,max_col),     # 从 B 列开始读取（跳过第一列）
    sheet_name=sheetname
    ).drop(0)   #读取表格
    oall=table.count().sum()

    cns=set()
    for index,content in table.items():
        goodsname=index
        for cn in content.values:
            if pd.isna(cn):
                continue
            else:
                cns.add(cn)
    return cns


name='App警察组吧唧（已到齐）'

#设置囤货
storages=['萌','dock','黄油']
storage=storages[1]

file_path = rf"./旧排发表/{storage}/{name}.xlsx"

filename=os.path.basename(file_path).split('.xlsx')[0]  #获取文件名/系列
obj_path=f'./colors/{name}.xlsx'   #生成目标地址

mainmodes=['shopping','schedule']
mainmode=mainmodes[1]
shoppingsheets=["8.20land"]
schedulesheets=[]

wb=load_workbook(file_path)
cns=set()

for sheetname in wb.sheetnames:
    cns.clear()
    if mainmode=='shopping':
        if sheetname in schedulesheets:
            cns=get_cns_schedule(name,file_path,wb,sheetname)
            colors=get_cell_colors(wb,cns,sheetname)
            save_or_append_to_excel(colors,obj_path,sheetname)
        else:
            continue
    elif mainmode=='schedule':
        if sheetname in shoppingsheets:
            continue
        else:
            cns=get_cns_schedule(name,file_path,wb,sheetname)
            colors=get_cell_colors(wb,cns,sheetname)
            save_or_append_to_excel(colors,obj_path,sheetname)

    




