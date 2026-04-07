#只做扫街模式
import pandas as pd
import os
from openpyxl import load_workbook
import numpy as np


name='7.13、15-18 扫街+7.28扫'

#设置囤货
storages=['萌','dock','黄油']
storage=storages[2]
        
file_path = rf"./排发表/{storage}/{name}.xlsx"

filename=os.path.basename(file_path).split('.xlsx')[0]  #获取文件名/系列
obj_path=f'./导入表/{storage}/{filename}'   #生成目标地址


wb=load_workbook(file_path)
for sheetname in wb.sheetnames:
    print("{}/{} processing.\n".format(name,sheetname))
    sheet=wb[sheetname] #遍历sheets并打开对应sheet
    max_col = sheet.max_column  #获取最大列数
    '''
    sheetname=pd.read_excel(file_path,header=None,nrows=1,sheet_name=sheetname)
    sheetname=sheetname.iloc[0,0].split('【')[-1].split('】')[0]
    '''
    table = pd.read_excel(
    file_path,
    skiprows=1,        # 跳过第一行
    sheet_name=sheetname
    )
    #统计扫街部分
    shopping_list=[]
    for index,contents in table.iterrows():
        if pd.isna(contents['cn']):
            pass
        else:
            cn=contents['cn']
        try:
            goodsname=contents['种类']
        except:
            goodsname=contents['谷名']
        shopping_list.append([cn,filename,sheetname,goodsname,1,'',storage,'不可排发'])

    shopping_list=pd.DataFrame(shopping_list,columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #重复项合并求和
    shopping_list=shopping_list.groupby(['cn','系列','表名','谷名','肾/国际状态','囤货','排发状态'],as_index=False).sum().reindex(columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #新建系列文件夹
    if not os.path.exists(obj_path):
        os.makedirs(obj_path,exist_ok=True)
    #按照sheetname分别存储
    shopping_list.to_excel(os.path.join(obj_path,f'{sheetname}.xlsx'),sheet_name=sheetname,header=True,index=False)
    print("{}/{} completed.\n".format(name,sheetname))
    