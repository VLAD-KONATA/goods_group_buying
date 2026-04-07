from openpyxl import load_workbook
import os

# 原 Excel 文件路径
input_file = "黄油处排发表.xlsx"
output_dir = "旧排发表"  # 输出目录

# 创建输出目录（如果不存在）
os.makedirs(output_dir, exist_ok=True)

# 加载原 Excel 文件（保留格式）
wb = load_workbook(input_file)

# 遍历每个 Sheet
for sheet_name in wb.sheetnames:
    # 创建新工作簿
    new_wb = load_workbook(input_file)  # 重新加载原文件以复制格式
    new_sheet = new_wb[sheet_name]  # 只保留当前 Sheet
    
    # 删除其他 Sheet
    for sheet in new_wb.sheetnames:
        if sheet != sheet_name:
            del new_wb[sheet]
    
    # 保存为单独的文件
    output_file = os.path.join(output_dir, f"{sheet_name}.xlsx")
    new_wb.save(output_file)
    print(f"已保存: {output_file}")

print("所有 Sheet 拆分完成！")