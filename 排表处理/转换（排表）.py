#只做排表模式
import pandas as pd
import os
from openpyxl import load_workbook
import numpy as np
#=IF(C4="",B4,CONCAT(B4,"（",C4,"）"))
#如果盲抽cn和种类为两列的excel处理方式

name='M28脱出系列'

#设置囤货
storages=['萌','dock','黄油']
storage=storages[2]

file_path = rf"./排发表/{storage}/{name}.xlsx"

filename=os.path.basename(file_path).split('.xlsx')[0]  #获取文件名/系列
obj_path=f'./导入表/{storage}/{filename}'   #生成目标地址

#筛选盲抽部分的正则表达式关键词
keywords=["盲抽"]
#keywords=["盲抽","随机","扭蛋","饮品卡","台北","高雄","台中"]



for i in range(len(keywords)):
    if i==0:
        keyline=keywords[i]
    else:
        keyline+='|'+keywords[i]    
        
wb=load_workbook(file_path)
for sheetname in wb.sheetnames:
    print("{}/{} processing.\n".format(name,sheetname))
    sheet=wb[sheetname] #遍历sheets并打开对应sheet
    max_col = sheet.max_column  #获取最大列数
    '''
    sheetname=pd.read_excel(file_path,header=None,nrows=1,sheet_name=sheetname)
    sheetname=sheetname.iloc[0,0].split('【')[-1].split('】')[0]
    '''
    table = pd.read_excel(
    file_path,
    skiprows=1,        # 跳过第一行
    usecols=range(1,max_col),     # 从 B 列开始读取（跳过第一列）
    sheet_name=sheetname
    ).drop(0)   #读取表格
    
    #通过正则筛选盲抽/非盲抽部分
    table_blind=table.loc[:, table.columns.str.contains(keyline,regex=True)]    
    table_unblind = table.loc[:, ~table.columns.str.contains(keyline,regex=True)]
    
    #统计非盲抽部分
    unblind_list=[]
    for index,content in table_unblind.items():
        goodsname=index
        for cn in content.values:
            if pd.isna(cn):
                continue
            else:
                unblind_list.append([cn,filename,sheetname,goodsname,1,'',storage,'不可排发'])
    unblind_list=pd.DataFrame(unblind_list,columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #重复项合并求和
    unblind_list=unblind_list.groupby(['cn','系列','表名','谷名','肾/国际状态','囤货','排发状态'],as_index=False).sum().reindex(columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    
    #统计盲抽部分
    blind_list=[]
    for index,contents in table_blind.items():
        goodsname=index
        for content in contents.values:
            if pd.isna(content):
                continue
            else:
                if "（" not in str(content) and "(" not in str(content):  #如果盲抽没有打开确认
                    cn=content
                    blind_list.append([cn,filename,sheetname,goodsname,1,'',storage,'不可排发'])
                else:
                    if content.split('（')[0]==content: #如果括号是英文括号
                        cn=content.split('(')[0]
                        chara=content.split('(')[-1].split(')')[0]
                    else:   #如果括号是中文括号
                        cn=content.split('（')[0]
                        chara=content.split('（')[-1].split('）')[0]
                        blind_list.append([cn,filename,sheetname,goodsname+chara,1,'',storage,'不可排发'])
    blind_list=pd.DataFrame(blind_list,columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #重复项合并求和
    blind_list=blind_list.groupby(['cn','系列','表名','谷名','肾/国际状态','囤货','排发状态'],as_index=False).sum().reindex(columns=['cn','系列','表名','谷名','数量','肾/国际状态','囤货','排发状态'])
    #合并盲抽/非盲抽部分
    final_list=pd.concat([blind_list,unblind_list], axis=0, ignore_index=True)
    #新建系列文件夹
    if not os.path.exists(obj_path):
        os.makedirs(obj_path,exist_ok=True)
    #按照sheetname分别存储
    final_list.to_excel(os.path.join(obj_path,f'{sheetname}.xlsx'),sheet_name=sheetname,header=True,index=False)
    print("{}/{} completed.\n".format(name,sheetname))

    